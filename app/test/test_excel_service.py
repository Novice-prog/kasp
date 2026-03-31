from app.services.excel_service import create_excel, EXCEL_CELL_CHAR_LIMIT
import os
from openpyxl import load_workbook

def test_create_excel(tmp_path):

    stats = {
        "житель": {
            "total": 3,
            "per_line": {0: 1, 1: 2}
        }
    }

    file_path = tmp_path / "result.xlsx"

    create_excel(stats, str(file_path))

    assert os.path.exists(file_path)

    wb = load_workbook(str(file_path))
    ws = wb.active

    assert ws["A1"].value == "Слово"
    assert ws["B1"].value == "Всего"
    assert ws["C1"].value == "По строкам"
    assert ws["A2"].value == "житель"
    assert ws["B2"].value == 3
    assert ws["C2"].value == "1,2"


def test_create_excel_truncates_very_long_cell_value(tmp_path):
    long_per_line = {i: 0 for i in range(40000)}
    long_per_line[0] = 1

    stats = {
        "житель": {
            "total": 1,
            "per_line": long_per_line
        }
    }

    file_path = tmp_path / "long.xlsx"
    create_excel(stats, str(file_path))

    wb = load_workbook(str(file_path))
    ws = wb.active
    value = ws["C2"].value

    assert len(value) <= EXCEL_CELL_CHAR_LIMIT
    assert value.endswith("...[TRUNCATED]")

